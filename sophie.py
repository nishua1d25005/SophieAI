#!/usr/bin/env python3
"""
SophieAI v2.0 — single-file, production-oriented assistant.

Features:
- Async event loop
- Secure config via environment variables
- Wake-word detection by streaming speech -> keyword matching
- Text & voice modes (switchable)
- OpenAI integration (uses OPENAI_API_KEY from env)
- Safe Excel operations via a constrained API (openpyxl)
- Google search + simple summary (optional)
- Memory saved to JSON with rotation
- Clean separation of responsibilities inside a single file

Dependencies:
pip install openai speechrecognition pyttsx3 googletrans==4.0.0-rc1 requests beautifulsoup4 openpyxl aiohttp

IMPORTANT: Set environment variables:
- OPENAI_API_KEY (required to use LLM features)
- SOPHIE_MEMORY_FILE (optional, default ./memory.json)
- OPENAI_MODEL (optional, default: gpt-3.5-turbo)
"""

import os
import sys
import asyncio
import json
import time
import datetime
import logging
from typing import Optional, Dict, Any, List

# Speech libraries
import speech_recognition as sr
import pyttsx3

# AI & web
import openai
import requests
from bs4 import BeautifulSoup
from googletrans import Translator

# Excel
import openpyxl

# -------------------------
# Configuration & Logging
# -------------------------
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("SophieAI")

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-3.5-turbo")  # change if you want
MEMORY_FILE = os.getenv("SOPHIE_MEMORY_FILE", "./memory.json")
MEMORY_MAX_ITEMS = 200

if not OPENAI_API_KEY:
    log.warning("OPENAI_API_KEY not set — GPT features will be disabled until you set it.")
else:
    openai.api_key = OPENAI_API_KEY

# -------------------------
# Utilities
# -------------------------
def safe_load_json(path: str, default: Any):
    if not os.path.exists(path):
        return default
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        log.error("Failed to load JSON (%s): %s", path, e)
        return default

def safe_save_json(path: str, obj: Any):
    tmp = path + ".tmp"
    try:
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(obj, f, indent=2, ensure_ascii=False)
        os.replace(tmp, path)
    except Exception as e:
        log.exception("Failed to save JSON: %s", e)

# -------------------------
# Memory manager
# -------------------------
class Memory:
    def __init__(self, path: str, max_items: int = MEMORY_MAX_ITEMS):
        self.path = path
        self.max_items = max_items
        self.data = safe_load_json(path, {"conversations": [], "last_topic": None})

    def append(self, user: str, assistant: str):
        self.data.setdefault("conversations", [])
        entry = {"ts": datetime.datetime.utcnow().isoformat() + "Z", "user": user, "assistant": assistant}
        self.data["conversations"].append(entry)
        if len(self.data["conversations"]) > self.max_items:
            self.data["conversations"] = self.data["conversations"][-self.max_items:]
        self.data["last_topic"] = user
        safe_save_json(self.path, self.data)

    def get_last_topic(self):
        return self.data.get("last_topic")

# -------------------------
# TTS (pyttsx3)
# -------------------------
class TTS:
    def __init__(self, voice_index: Optional[int] = None, rate: Optional[int] = 180):
        self.engine = pyttsx3.init()
        if voice_index is not None:
            try:
                voices = self.engine.getProperty("voices")
                if 0 <= voice_index < len(voices):
                    self.engine.setProperty("voice", voices[voice_index].id)
            except Exception:
                pass
        if rate is not None:
            try:
                self.engine.setProperty("rate", rate)
            except Exception:
                pass

    def speak(self, text: str):
        if not text:
            return
        log.info("Sophie: %s", text)
        try:
            self.engine.say(text)
            self.engine.runAndWait()
        except Exception as e:
            log.exception("TTS error: %s", e)

# -------------------------
# Speech recognition (sync wrapper)
# -------------------------
class SpeechListener:
    def __init__(self, energy_threshold: int = 300, pause_threshold: float = 0.5, language: str = "en-IN"):
        self.recognizer = sr.Recognizer()
        self.recognizer.energy_threshold = energy_threshold
        self.recognizer.pause_threshold = pause_threshold
        self.language = language
        self.microphone = sr.Microphone()

    def listen_once(self, timeout: Optional[float] = None, phrase_time_limit: Optional[float] = 8.0) -> str:
        """Listen once and return text. Non-blocking wrapper would need threading — kept simple here."""
        with self.microphone as source:
            self.recognizer.adjust_for_ambient_noise(source, duration=0.7)
            log.debug("Listening (timeout=%s, limit=%s)...", timeout, phrase_time_limit)
            audio = self.recognizer.listen(source, timeout=timeout, phrase_time_limit=phrase_time_limit)
        try:
            text = self.recognizer.recognize_google(audio, language=self.language)
            log.info("User said: %s", text)
            return text
        except sr.UnknownValueError:
            return ""
        except sr.RequestError as e:
            log.error("Speech recognition request failed: %s", e)
            return ""

# -------------------------
# OpenAI Chat helper (safe wrapper)
# -------------------------
async def chat_with_openai(prompt: str, model: str = OPENAI_MODEL, max_tokens: int = 350) -> str:
    if not OPENAI_API_KEY:
        return "OpenAI API key not configured."
    loop = asyncio.get_event_loop()
    try:
        # Use the synchronous client inside threadpool to avoid blocking
        def call_api():
            resp = openai.ChatCompletion.create(
                model=model,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=max_tokens,
                temperature=0.2,
            )
            return resp
        response = await loop.run_in_executor(None, call_api)
        text = response["choices"][0]["message"]["content"].strip()
        return text
    except Exception as e:
        log.exception("OpenAI error: %s", e)
        return f"Error contacting OpenAI: {e}"

# -------------------------
# Simple safe Excel API (explicit allowed ops only)
# -------------------------
ALLOWED_EXCEL_OPS = {"create_workbook", "write_cell", "append_row", "read_cell", "list_sheets"}

def perform_excel_task(spec: Dict[str, Any]) -> Dict[str, Any]:
    """
    spec example:
    {
       "op": "write_cell",
       "file": "data.xlsx",
       "sheet": "Sheet1",
       "cell": "A1",
       "value": "hello"
    }
    """
    op = spec.get("op")
    if op not in ALLOWED_EXCEL_OPS:
        return {"error": "operation_not_allowed"}

    file = spec.get("file")
    if not file:
        return {"error": "file_required"}

    try:
        if op == "create_workbook":
            wb = openpyxl.Workbook()
            wb.save(file)
            return {"ok": True, "message": f"Created {file}"}

        if op in ("write_cell", "append_row", "read_cell", "list_sheets"):
            if not os.path.exists(file):
                return {"error": "file_not_found"}
            wb = openpyxl.load_workbook(file)
            sheet_name = spec.get("sheet", wb.sheetnames[0])

            if op == "write_cell":
                ws = wb[sheet_name]
                ws[spec["cell"]] = spec["value"]
                wb.save(file)
                return {"ok": True}

            if op == "append_row":
                ws = wb[sheet_name]
                ws.append(spec.get("row", []))
                wb.save(file)
                return {"ok": True}

            if op == "read_cell":
                ws = wb[sheet_name]
                return {"ok": True, "value": ws[spec["cell"]].value}

            if op == "list_sheets":
                return {"ok": True, "sheets": wb.sheetnames}

    except Exception as e:
        log.exception("Excel operation error: %s", e)
        return {"error": str(e)}

# -------------------------
# Web search + summary (very small)
# -------------------------
def google_search_and_summary(query: str, num_results: int = 3) -> Dict[str, Any]:
    """
    Basic google search using the 'requests' approach to 'google' is fragile.
    This function uses the free 'google search' approach by hitting the 'ngram' endpoints is not reliable.
    For production, use an official search API (Custom Search JSON API, SerpAPI, etc).
    Here we attempt a minimal approach with duckduckgo html scrapes (lightweight).
    """
    try:
        url = f"https://duckduckgo.com/html/?q={requests.utils.quote(query)}"
        headers = {"User-Agent": "Mozilla/5.0"}
        r = requests.get(url, headers=headers, timeout=6)
        soup = BeautifulSoup(r.text, "html.parser")
        results = []
        for a in soup.select(".result__a")[:num_results]:
            href = a.get("href")
            title = a.get_text().strip()
            results.append({"title": title, "href": href})
        return {"ok": True, "results": results}
    except Exception as e:
        log.exception("Search error: %s", e)
        return {"ok": False, "error": str(e)}

def fetch_page_summary(url: str, max_paragraphs: int = 3) -> str:
    try:
        headers = {"User-Agent": "Mozilla/5.0"}
        r = requests.get(url, headers=headers, timeout=6)
        soup = BeautifulSoup(r.content, "html.parser")
        paragraphs = soup.find_all("p")
        if not paragraphs:
            return "No textual summary found."
        text = " ".join(p.get_text().strip() for p in paragraphs[:max_paragraphs])
        return text[:1500]  # keep it bounded
    except Exception as e:
        log.exception("Summary fetch error: %s", e)
        return f"Could not fetch summary: {e}"

# -------------------------
# Main Sophie class — orchestrates
# -------------------------
class Sophie:
    def __init__(self, mode: str = "both", wake_word: str = "sophie"):
        self.mode = mode  # "voice", "text", "both"
        self.wake_word = wake_word.lower()
        self.memory = Memory(MEMORY_FILE)
        self.tts = TTS()
        self.listener = SpeechListener()
        self.translator = Translator()

    async def start(self):
        log.info("Sophie starting in %s mode (wake word='%s')", self.mode, self.wake_word)
        if self.mode in ("voice", "both"):
            await self.loop_voice()
        else:
            await self.loop_text()

    async def loop_text(self):
        while True:
            try:
                user = input("You: ").strip()
            except EOFError:
                break
            if not user:
                continue
            resp = await self.handle_user_input(user, via_voice=False)
            print("Sophie:", resp)
            self.tts.speak(resp)

    async def loop_voice(self):
        """
        Simple voice loop: listens in short chunks, checks for wake word if not active,
        then captures command and processes it. This is intentionally conservative to avoid
        false positives.
        """
        log.info("Entering voice loop. Say the wake word ('%s') to activate.", self.wake_word)
        while True:
            # 1) Listen a short phrase
            text = self.listener.listen_once(timeout=5, phrase_time_limit=6)
            if not text:
                # no speech recognized
                await asyncio.sleep(0.1)
                continue
            text = text.strip()
            # 2) Check for wake word
            if self.wake_word in text.lower():
                self.tts.speak("Yes sir, I'm listening.")
                # Capture a longer command now
                cmd = self.listener.listen_once(timeout=8, phrase_time_limit=12)
                if not cmd:
                    self.tts.speak("I didn't catch that. Say again.")
                    continue
                cmd = cmd.strip()
                # Process
                response = await self.handle_user_input(cmd, via_voice=True)
                self.tts.speak(response)
            else:
                # If mode == both, allow text triggers or pass
                if self.mode == "both":
                    # treat as ambient utterance; optionally process short commands
                    if len(text.split()) <= 3 and any(k in text.lower() for k in ["time", "date", "hello", "hi"]):
                        response = await self.handle_user_input(text, via_voice=True)
                        self.tts.speak(response)
                # else ignore until wake word
            await asyncio.sleep(0.05)

    async def handle_user_input(self, user_input: str, via_voice: bool = False) -> str:
        """
        Central command dispatcher. Keep it readable and extensible.
        """
        # Normalize
        text = user_input.strip()
        if not text:
            return "I didn't hear anything."

        # Simple exits
        if text.lower() in ("exit", "quit", "goodbye", "bye", "stop"):
            return "Goodbye. Take care."

        # Language detection (very simple heuristic)
        if any(ch in text for ch in "अआइईउऊएऐओऔनम"):
            # Hindi-ish -> translate to English for processing
            try:
                text = self.translator.translate(text, src="hi", dest="en").text
                log.debug("Translated input -> %s", text)
            except Exception:
                pass

        # COMMAND: Excel safe tasks (JSON-ish commands)
        if text.lower().startswith("excel"):
            # Expect: excel: {"op": "...", ...}
            try:
                jsonpart = text.partition(":")[2].strip()
                spec = json.loads(jsonpart)
                result = perform_excel_task(spec)
                return json.dumps(result, ensure_ascii=False)
            except Exception as e:
                log.exception("Excel parsing error")
                return "I couldn't parse the Excel command. Use: excel: {json-spec}"

        # COMMAND: open app (limited, safe)
        if any(kw in text.lower() for kw in ("open notepad", "open calculator", "open chrome")):
            if "notepad" in text.lower():
                if sys.platform.startswith("win"):
                    os.system("start notepad")
                    return "Opened Notepad."
                else:
                    return "Notepad command works only on Windows."
            if "calculator" in text.lower():
                if sys.platform.startswith("win"):
                    os.system("start calc")
                    return "Opened Calculator."
                else:
                    return "Calculator command platform-dependent."
            if "chrome" in text.lower():
                # open default browser to Google homepage
                import webbrowser
                webbrowser.open("https://www.google.com")
                return "Opened browser."

        # COMMAND: time/date
        if any(kw in text.lower() for kw in ("time", "what time", "date", "what date")):
            now = datetime.datetime.now()
            if "time" in text.lower():
                return f"The time is {now.strftime('%H:%M:%S')}."
            else:
                return f"Today is {now.strftime('%A, %d %B %Y')}."

        # COMMAND: news (simple search+summary)
        if "news" in text.lower():
            q = text.replace("news", "").strip() or "latest news"
            sres = google_search_and_summary(q, num_results=2)
            if not sres.get("ok"):
                return "I couldn't fetch news right now."
            lines = [f"{i+1}. {r['title']}" for i, r in enumerate(sres.get("results", []))]
            return "Here are top results: " + " | ".join(lines)

        # COMMAND: web search
        if text.lower().startswith("search ") or text.lower().startswith("google "):
            q = text.split(" ", 1)[1]
            sres = google_search_and_summary(q, num_results=3)
            if not sres.get("ok"):
                return "Search failed."
            results = sres.get("results", [])
            if not results:
                return "No results found."
            # Fetch summary for first result (safe bounded)
            first = results[0]
            summary = fetch_page_summary(first.get("href") or "", max_paragraphs=2)
            return f"{first.get('title')} — {summary}"

        # COMMAND: ask LLM (prefix gpt:)
        if text.lower().startswith("gpt:") or text.lower().startswith("ask:"):
            prompt = text.partition(":")[2].strip()
            if not prompt:
                return "Provide a prompt after 'gpt:'"
            answer = await chat_with_openai(prompt)
            # Save conversation
            self.memory.append(user_input, answer)
            return answer

        # Fallback: intent classification by a small rule set or LLM
        if len(text) < 200 and sum(len(w) for w in text.split()) < 100:
            # ask LLM for a short intent classification (low cost)
            prompt = f"""You are a concise intent classifier. Provide a one-line summary intent and one-line action for this user input:
User Input: \"{text}\"
Respond in JSON: {{ "intent": "...", "action": "..." }}"""
            resp = await chat_with_openai(prompt, max_tokens=120)
            # try to parse JSON from the response
            try:
                parsed = json.loads(resp)
                intent = parsed.get("intent", "")
                action = parsed.get("action", "")
                # very small set of intent actions we can execute
                if action.lower().startswith("open "):
                    # let user know
                    return f"Intent: {intent}. Action recommended: {action}"
                # Save memory
                self.memory.append(user_input, resp)
                return f"Intent detected: {intent}. Suggestion: {action}"
            except Exception:
                # If parsing fails, return the LLM answer as-is
                self.memory.append(user_input, resp)
                return resp

        # Last fallback: small LLM completion
        answer = await chat_with_openai(f"Answer concisely: {text}", max_tokens=250)
        self.memory.append(user_input, answer)
        return answer

# -------------------------
# CLI / Entrypoint
# -------------------------
def parse_args():
    import argparse
    p = argparse.ArgumentParser(prog="sophie_v2", description="SophieAI v2 assistant")
    p.add_argument("--mode", choices=["text", "voice", "both"], default="both", help="Interaction mode")
    p.add_argument("--wake", default="sophie", help="Wake word")
    return p.parse_args()

async def main():
    args = parse_args()
    sophie = Sophie(mode=args.mode, wake_word=args.wake)
    try:
        await sophie.start()
    except KeyboardInterrupt:
        log.info("Shutting down Sophie (KeyboardInterrupt).")
    except Exception:
        log.exception("Unhandled exception in main loop.")

if __name__ == "__main__":
    asyncio.run(main())
